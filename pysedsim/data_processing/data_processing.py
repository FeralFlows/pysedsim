# -*- coding: utf-8 -*-

'''

Module imports and exports data inputs and outputs to and from PySedSim.

Contains numerous functions to do this.

'''

# Import relevant libraries
from __future__ import division
from openpyxl import load_workbook  # Purpose: load excel workbooks
from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
import pandas as pd
import os  # allows export of csv file by joining intended file name with the destination folder's path.
import calendar  # for using the .isleap function to determine whether current year is a leap year.
import numpy as np
import platform  # Purpose: tells you whether your computer/cluster is windows or linux.
import csv
import itertools
import logging

def Determine_Num_Scenarios(file_name='PySedSim_Input_Specifications.csv'):
    '''

    Purpose: Determine basic information about the simulation that either was already, or will be, conducted,
    such as simulation name, names of elements and variables simulated, location of files, and number of elements
    simulated.

    :param file_name: String, name of the top-level input file (e.g., 'PySedSim_Input_Specifications.csv')
    :return:
    1. num_scenarios: the number of scenarios (each possibly w/ a different network, but with same number of
    realizations to be performed). If running PySedSim in parallel, each processor's workload is increased by a
    factor of num_scenarios.
    2. simulation_titles_list: list of scenario names (simulation titles)
    3. imported_specs: basic simulation specifications (taken from imported_data)
    4. main_input_files_dir: location on computer of input files to run simulation (or already used to run simulation)
    5. main_output_file_dir: location on computer to store output files (or where output files have already been
    stored)

    '''

    imported_data = list(csv.reader(file(file_name, 'rb')))  # Import entire file
    imported_specs = [column[1] for column in imported_data[1:6]]
    simulation_titles_list = imported_specs[0]  # e.g., "Sambor Original"  # "Alternative with Flushing"  # "Alternative with NO Flushing"
    simulation_titles_list = simulation_titles_list.split(', ')  # list of scenarios to be simulated
    num_scenarios = len(simulation_titles_list)  # number of scenarios in above list
    if num_scenarios == 0:
        logging.critical("User must specify scenario name(s) in top-level input file (e.g., "
                      "PySedSim_Input_Specifications.csv")
    Monte_Carlo_Parameters_File_List = imported_specs[3]
    Monte_Carlo_Parameters_File_List = Monte_Carlo_Parameters_File_List.split(', ')

    # Name/location of Monte Carlo parameters sheet. If no sheet is named, then user intends to supply all files. If a sheet is
    # named, the model will generate the MC inputs.
    if len(Monte_Carlo_Parameters_File_List) == 0:
        # User will provide files with all stochastic inputs specified, as no input file with parameter ranges and distributions has
        # been specified.
        external_mc_data = 1
    else:
        external_mc_data = 0
        if num_scenarios > 1 and len(Monte_Carlo_Parameters_File_List) <= 1:
            # User did not specify enough MC input files for corresponding number of scenarios. Just use the same
            # MC file for all scenarios.
            for i in range(int(num_scenarios-len(Monte_Carlo_Parameters_File_List))):
                Monte_Carlo_Parameters_File_List.append(Monte_Carlo_Parameters_File_List[0])
    main_input_files_dir = imported_specs[1]
    main_output_file_dir = imported_specs[4]
    return num_scenarios, simulation_titles_list, imported_specs, main_input_files_dir, main_output_file_dir, \
           Monte_Carlo_Parameters_File_List, external_mc_data

def Import_Simulation_Preferences(imported_specs, simulation_title, main_input_files_dir, start_stop = None,
                                  export = None, re_eval = None, Monte_Carlo_Parameters_File_Name = None,
                                  external_mc_data = 0):

    # Purpose: To import the user's simulation preferences from the input .xlsx sheet (e.g., "Input_File.xlsx").
    # Inputs:
    # (1) imported_specs: produced by Determine_Num_Scenarios() function
    # (2) simulation_title: the scenario (1 taken from list of titles in simulation_titles_list) currently being simulated
    # (3) start_stop: (optional) a list containing the range of monte carlo inputs over which this processor will loop (e.g., [5, 13])
    # (4) export: (optional) contains a string that is either 'Yes' or 'No', indicating whether output from PySedSim
    # simulation should be exported. This value should be set = 'No' as an input in an optimization setting,
    # as it will become very expensive to continually export output to .csv files in this case.
    # (5) re_eval_vars: List of variable names (strings) for which outputs should be exported, if listed in the
    # "Reevaluation" worksheet of the input file. Used for deterministic or stochastic reevaluation of particular
    # operating policies.

    export_file_type = 'csv'  # Currently only choice is .csv. In future may add .xls/.xlsx support.
    # if you select 'No', the program will attempt to store and export all time series variables, # so be sure your RAM is capable of
    # handling this data load.

    # Get operator for changing directory based on operating system.
    os_fold_op = Op_Sys_Folder_Operator()

    Input_Data_File = Load_Input_File(simulation_title, main_input_files_dir, imported_specs,
                                      os_fold_op=os_fold_op)

    SedSim_File_Input_Sheet = Input_Data_File['Simulation Specifications']
    # Simulation Dates and Durations
    #start_date = datetime(int(SedSim_File_Input_Sheet['B3'].value),int(SedSim_File_Input_Sheet['C3'].value),
    #                      int(SedSim_File_Input_Sheet['D3'].value))
    #end_date = datetime(int(SedSim_File_Input_Sheet['B4'].value), int(SedSim_File_Input_Sheet['C4'].value),
    #                      int(SedSim_File_Input_Sheet['D4'].value))
    #Sim_Dur = int((end_date - start_date).days + 1)

    start_date = SedSim_File_Input_Sheet['B3'].value
    end_date = SedSim_File_Input_Sheet['B4'].value
    Sim_Dur = (end_date - start_date).days + 1
    T = Sim_Dur

    # Time series state variables to be exported
    if export is None:
        # User is not doing an optimization, so read simulation export preferences from input file.
        [var_sub_list, element_export_list, export_data] = Export_Preferences(Input_Data_File)

        # Add variables to var_sub_list and element_export_list if user is conducting a policy re-evaluation,
        # with preferences specified in the "Reevaluation" worksheet of the input file.
        if re_eval is not None:
            for key in re_eval:
                if re_eval[key]['State Variable Name'] not in var_sub_list:
                    var_sub_list.append(re_eval[key]['State Variable Name'])
                if re_eval[key]['Location'] not in element_export_list:
                    element_export_list.append(re_eval[key]['Location'])
    else:
        # User is doing an optimization. Do not print output to .csv files.
        var_sub_list = None
        element_export_list = None
        export_data = export

    # If simulation will be stochastic (monte-carlo style), import relevant data and parameters
    # First, establish default values for some parameters, whose values are included in "return" and thus must be at least initialized.

    if re_eval is not None:
        # User is conducting a reevaluation of operating policies.
        if re_eval[re_eval.keys()[0]]['sim_type'] == 'Stochastic':
            Stochastic_Sim = 1
            Num_Realizations = re_eval[re_eval.keys()[0]]['num_reevals']
        else:
            Stochastic_Sim = 0
            Num_Realizations = 1
    else:
        if SedSim_File_Input_Sheet['B5'].value == 'Stochastic':
            Stochastic_Sim = 1
            Num_Realizations = Input_Data_File['Simulation Specifications']['B6'].value
        else:
            Stochastic_Sim = 0
            Num_Realizations = 1

    if Stochastic_Sim == 1:
        # Stochastic simulation will be performed.
        if start_stop is not None:
            start = start_stop[0]
            stop = start_stop[1]
            Num_Realizations = stop - start
        else:
            start = 0
            stop = Num_Realizations
            start_stop = [start]
            start_stop.append(stop)
        # Create a list of column names (i.e., Realization1, Realization2, ..., Realization100)
        Col_Names = ['Realization' + str(x+1) for x in range(start, stop)]  # Label cols as "Realization1, ..., RealizationN"

        simulation_dates = pd.date_range(start_date, end_date)  # will serve as DF index for the time series data.
        simulation_dates_no_leap = pd.date_range(start_date, start_date + timedelta(
            (simulation_dates[-1].year - simulation_dates[0].year + 1) * 365 - 1))
    else:
        # Deterministic simulation is the default:
        Col_Names = ['Realization1']
        simulation_dates = pd.date_range(start_date, start_date + timedelta(Sim_Dur - 1))  # will serve as DF index for time series data.
        simulation_dates_no_leap = simulation_dates

    Sampled_Parameter_Dict = {}  # This will be populated later if the simulation is stochastic.
    Synthetic_Inflow_dataframe_name_LIST = None  # This will be populated later if the simulation is stochastic.
    Synthetic_Inflows_dictionary = {}  # This will be populated later if the simulation is stochastic.

    return export_data, export_file_type, var_sub_list, element_export_list, \
           Input_Data_File, T, Sim_Dur, Stochastic_Sim, Num_Realizations, simulation_dates, \
           simulation_dates_no_leap, Col_Names, Monte_Carlo_Parameters_File_Name, external_mc_data, simulation_title,\
           start_stop, Sampled_Parameter_Dict, Synthetic_Inflow_dataframe_name_LIST, Synthetic_Inflows_dictionary


def Load_Input_File(simulation_title, main_input_files_dir, imported_specs, os_fold_op=None):

    '''

    Purpose: To create the input data .xlsx file for purposes of importing data in particular worksheets in the file.
    Uses openpyxl to do this, and can only work with .xlsx type files.

    :param simulation_title: Name of simulation being run, as named in input specification .csv sheet. String.
    :param main_input_files_dir: Name of directory on computer where top level input specifications .csv file is stored.
    :param os_fold_op: Result of running os_fold_op() method. Determines forward or backslash
    requirement based on operating system.
    :param imported_specs: A list determined from the data_processing module in the determine_num_scenarios() method.
    Contains basic data provided in input specification .csv sheet.
    :return: the openpyxl object .xlsx file and its name
    '''

    if os_fold_op is None:
        os_fold_op = Op_Sys_Folder_Operator()

    if len(imported_specs[2]) == 0:
        # Simulation input file name not specified. This indicates title is "PySedSim_Input_File - Simulation_Title.xlsx"
        Input_Data_File_Name = "PySedSim_Input_File - " + simulation_title + '.xlsx'
    else:
        Input_Data_File_Name = imported_specs[2]  # Use user-provided value
    Input_Data_File = load_workbook(filename=main_input_files_dir + os_fold_op + Input_Data_File_Name, read_only = False,
                                    data_only=True)
    return Input_Data_File

def Export_Preferences(Input_Data_File):

    '''

    :param Input_Data_File: The .xlsx input data file (openpyxl object). Can be externally created or using
    Load_Input_File() function in this module.
    :return: var_sub_list, element_export_list, export_data

    1. var_sub_list = list of state variables to be exported for elements in element_export_list. Default=None if
    user has no preferences, meaning later it will include all state variables for corresponding element

    2. element_export_list= list of system elements for which data should be exported for variables in var_sub_list.
    Default = None if user has no preferences, meaning later it will include all system elements.

    3. export_data = 'Yes' or "No'. Default = 'Yes'.

    '''

    var_sub_list = None  # Default is NoneType. If not reset below, all variables will be exported.
    element_export_list = None  # Default is NoneType. If not reset below, all elements will be exported.
    export_data = 'Yes'  # Default is to export data for all system elements and relevant time series variables.
    if 'Export Preferences' in Input_Data_File.sheetnames:
        if Input_Data_File['Export Preferences']['B1'].value in ['No', 'no']:
            export_data = 'No'
        else:
            export_data = 'Yes'  # Default is to export simulation output.
            var_sub_list = Input_Data_File['Export Preferences']['B3'].value
            if var_sub_list is not None:
                var_sub_list = Input_Data_File['Export Preferences']['B3'].value.split(', ')  # Read in values
            else:
                pass  # User did not specify locations to export, so export all variables for location.
            element_export_list = Input_Data_File['Export Preferences']['B2'].value
            if element_export_list is not None:
                element_export_list = Input_Data_File['Export Preferences']['B2'].value.split(', ')  # Read in values
            else:
                pass  # User did not specify locations to export, so export all variables for location.
    else:
        pass  # All variables for export elements will be exported.

    return var_sub_list, element_export_list, export_data

def Monte_Carlo_Import(main_input_files_dir, Col_Names, element_stochastic_components, Element_List,
                       external_mc_data, Monte_Carlo_Parameters_File_Name, simulation_dates,
                       simulation_dates_no_leap, Num_Realizations, Sampled_Parameter_Dict, Synthetic_Inflow_dataframe_name_LIST,
                       Synthetic_Inflows_dictionary, start_stop = None, parallelize_import = None):

    # Purpose: To import user's monte carlo preferences from an input .xlsx sheet (e.g., "Monte_Carlo_Input_File.xlsx").

    # Inputs:
    # (2) main_input_files_dir (Specified by user)
    # (3) Col_Names (Created automatically in Import_Simulation_Preferences(), Stores list of state var names for each object instance).
    # (4) System_Object_List (same as SystemObjects["Ordered Simulation List"] created from System_Element_Creation())
    # (5) element_stochastic_components is both an input and an output. It gets modified here and dumped out.

    # Outputs:
    # (1) Parameter_Input_Dictionary
    # (2) element_stochastic_components

    # Define input parameters if Monte Carlo Simulation is to be done. Only create once before main simulation loop.
    # See parameter_sampling_monte_carlo.py for parameter listing and dictionary formatting requirements.
    os_fold_op = Op_Sys_Folder_Operator()

    # Establish leap year preferences
    leap_year_included = 0  # Options are 0 or 1
    if leap_year_included == 0:
        sim_dates = simulation_dates_no_leap
    else:
        sim_dates = simulation_dates

    # Define input parameters for stochastic simulation
    Parameter_Input_Dictionary = {}  # Define dictionary

    # Locate Monte-Carlo input file, if user wishes to generate own monte carlo values.
    if external_mc_data == 0:
        # User needs to generate MC parameter values. Import relevant preferences from Monte_Carlo_Inputs.xls
        Monte_Carlo_Specs_File = load_workbook(filename=main_input_files_dir + os_fold_op + Monte_Carlo_Parameters_File_Name,
                                               read_only = False, data_only=True)
        import_cols_dict = {'1': 2, '2': 1, '3': 4, '4': 3, '5': 3, '6': 'N/A', '7': 3, '8': 3, '9': 'N/A', '10': 3,
                            '11': 3, '12': 3, '13': 3, '14': 3}

        for locations in Element_List:
            # Determine what sheets to import data from for each element in 'Master' worksheet. Read in a string of sheet names,
            # and convert that to a list. If value is None, then the system element will not have information in any sheets.
            element_stochastic_components[locations] = Excel_Data_Import(locations, Monte_Carlo_Specs_File, 'Master', 1, 1,
                                                                         max_distinct_data_types=None, data_name_offset=None)
            Parameter_Input_Dictionary[locations] = {}  # Initialize nested sub-dict for each location.
            if element_stochastic_components[locations][0] is not None:
                if type(element_stochastic_components[locations][0]) in [str, unicode]:
                    # User has specified a list of numbers, which will be imported as one long string with commas
                    element_stochastic_components[locations] = element_stochastic_components[locations][0].split(', ')
                else:
                    # User specified only only one sheet number, which will be imported as a single number (long type)
                    element_stochastic_components[locations] = [str(element_stochastic_components[locations][0])]
            elif element_stochastic_components[locations][0] is None:
                element_stochastic_components[locations] = []

        # Loop through all system elements and store monte carlo parameter preferences so they can be generated.
        for locs in Element_List:
            for sheet in element_stochastic_components[locs]:
                Parameter_Input_Dictionary[locs][sheet] = {'Params': Excel_Data_Import(locs, Monte_Carlo_Specs_File, sheet, 1,
                                                                                       import_cols_dict[sheet],
                                                                                       max_distinct_data_types=None,
                                                                                       data_name_offset=None)}
    else:
        for locs in Element_List:
            Parameter_Input_Dictionary[locs] = {}

    # For stochastically generated incremental flows and sediment loads, loop through input file directory to see if any workbooks are
    # the same name as system junctions, in which case store those names and flow values.
    Synthetic_Inflow_Locations_List = []
    for locations in Element_List:
        file_stored = 0  # Binary, whether or not file specified for inflows for junction
        file_path_test = main_input_files_dir + os_fold_op
        try:
            if Parameter_Input_Dictionary[locations]['1']['Params'][0] is not None:
                file_stored = 1
                file_path = file_path_test + Parameter_Input_Dictionary[locations]['1']['Params'][0]  # use specified
                # input file name
                Synthetic_Inflow_Locations_List.append(locations)  # e.g., "Junction 1"
            else:
                file_stored = 1
                file_path = file_path_test + locations + "_Flow.csv"
                Synthetic_Inflow_Locations_List.append(locations)  # e.g., "Junction 1"
        except KeyError:
            # Default file name for flows is location_Flow.csv, where "location" is name of junction.
            file_path_test += locations + "_Flow.csv"
            if os.path.isfile(file_path_test):
                file_path = file_path_test
                file_stored = 1
                Synthetic_Inflow_Locations_List.append(locations)  # e.g., "Junction 1"

        # Read in any preferences regarding inclusion of leap years
        try:
            if Parameter_Input_Dictionary[locations]['1']['Params'][1] is not None:
                if Parameter_Input_Dictionary[locations]['1']['Params'][1] in ['Y', 'y', 'Yes', 'yes']:
                    leap_year_included = 1
                else:
                    leap_year_included = 0
            else:
                leap_year_included = 0  # Default
        except KeyError:
            leap_year_included = 0
        if leap_year_included == 0:
            sim_dates = simulation_dates_no_leap
        else:
            sim_dates = simulation_dates

        # Store preferences for this location
        if file_stored == 1:
            try:
                Parameter_Input_Dictionary[locations]['1']
            except KeyError:
                Parameter_Input_Dictionary[locations]['1'] = {}
            Parameter_Input_Dictionary[locations]['1']['Preferences'] = {
                'File Path': file_path,
                'Locations': [locations],
                'File Type': 'csv',
                'Leap Year Included': leap_year_included,
                'Date Range': simulation_dates,
                'Column Names': Col_Names}
                # Synthetic_Inflow_Locations_List
    # Call Monte Carlo parameter sampling function. Only need to do once if simulating on one computer. Must be done after
    # network connectivity has been defined.

    # Get leap year preferneces, as they may have been inadvertently overwritten for locations without incremental
    # flows.
    for locations in Element_List:
        try:
            if Parameter_Input_Dictionary[locations]['1']['Preferences']['Leap Year Included'] == 0:
                sim_dates = simulation_dates_no_leap
            else:
                sim_dates = simulation_dates
            break  # Found value, now exit, having stored preferences.
        except KeyError:
            pass
    if len(Synthetic_Inflow_Locations_List) > 0:
        [Sampled_Parameter_Dict, Synthetic_Inflow_dataframe_name_LIST, Synthetic_Inflows_dictionary] = Parameter_Sampling_Monte_Carlo(
            Num_Realizations, Parameter_Input_Dictionary, start_stop=start_stop, simulation_dates=sim_dates,
            parallelize_import=parallelize_import)
    else:
        Sampled_Parameter_Dict = Parameter_Sampling_Monte_Carlo(Num_Realizations, Parameter_Input_Dictionary, start_stop=start_stop, simulation_dates=sim_dates)
    return Parameter_Input_Dictionary, element_stochastic_components, Sampled_Parameter_Dict, Synthetic_Inflow_dataframe_name_LIST, \
           Synthetic_Inflows_dictionary

def Parameter_Sampling_Monte_Carlo(Num_Realizations, Parameter_Input_Dictionary, start_stop=None,
                                   simulation_dates=None, parallelize_import = None):
    '''
    Purpose: This function is designed to receive parameter input ranges and distributions, and will return sampled
    parameter sets to run in a Monte-Carlo Simulation mode.

    Inputs:

    :param Num_Realizations: Number of simulations to be run, each with a different set of parameters/forcings
    :param simulation_dates: a pandas date_range object that will serve as a dataframe index for the time series data.
    # (5) Sample Format: Parameter_Input_Dictionary['5'] = ['triangular', 50, 80, 100]
    :param Parameter_Input_Dictionary: indicating which parameters will be sampled, where parameters are given by
    the numbers below.

    # WARNINGS:
    # 1. Users are recommended to use save any .xls or .xlsx files as .csv files.
    # 2. For a csv file you MUST include at least one column heading (e.g., realization1). This will allow the dataframe to be interpreted properly and columns/indices set.
    # 3. Use a .xlsx file if you need to import synthetic flows as a dictionary from more than one location.
    # 4. If you use .csv, you can't parse columns, which means you need to specify a number of realizations in the input file that is equal to the number of equivalent columns contained in the CSV file. So you can't pluck 3 cols from the 500 realizations in a csv file.

    # Numbers for different parameters that can be sampled:
    # Note: Parameter Dictionary Keys may be any of the following numbers, which correspond to different inputs

    # 1. Hydrologic Inflows
    # 2. Sediment Load Inflows: mean = provide (a,b) if they are not going to be randomly sampled; variance = v; distribution from which to sample noise
    # 3. Annual Sediment Load: ASL
    # 4. Rating curve exponent: a
    # 5. Rating curve coefficient: b
    # 6. Trapping Efficiency Parameters: m, a, b, c, d
    # 7. Trapping Efficiency Curve (Brune) -- can be L, M or H
    # 8. Sediment density (kg/m^3)
    # 9. Entire E_Sed curve that describes what cumulative fraction of sediment is deposited below each elevation in the E-V-A curve. See option 10 below for a potentially simpler option
    # 10. Fraction of sediment that will be deposited within the active storage zone. So 1-value = amount cumulatively stored below this elevation in each time step. Must be 0<=value<=1.
    # 11. Bypass fraction
    # 12. Amount by which to shift churchill curve trapping up/down from its current value
    # 13. self.W_fb_coeff for Reservoir.Operating_Policy["Flushing"]. Coefficient used to define function that describes flushing channel bottom width. This value will be used in deterministic simulations. Stochastic simulations will externally specify this value.
    # 14. Functional form of Egg passage drawdown curve, which comes from hydraulic modeling.

    # Example of Parameter_Dictionary for Annual Sediment Load (Item # 3 below). Distribution must be the name of a numpy distribution, and parameters must be set in order in which the numpy distribution takes that parameter.

    Outputs:
    Sampled parameter value dictionary: Sampled_Parameter_Dict. Contains Num_Realizations of parameters for each parameter. Will be similar format to input dictionary with regard to key names.

    '''

    Sampled_Parameter_Dict = {}  # Initialize dictionary.
    os_fold_op = Op_Sys_Folder_Operator()
    Synthetic_Inflows_dictionary = {}  # Initialize
    Synthetic_Inflow_dataframe_name_LIST = {}  # Create dictionary that will store lists of names of data frame members, e.g., will store dict['Kratie'] = ['Realization1', 'Realization2', etc.]

    for locs in Parameter_Input_Dictionary:
        # Loop through locations and generate random values for each location/variable. First handle flow import separately.
        Sampled_Parameter_Dict[locs] = {}  # Init nested sub-dict
        for keys in Parameter_Input_Dictionary[locs]:
            if keys in ['1']:
                # Stochastic hydrologic inputs are specified
                # Import a DataFrame containing all stochastic hydrologic and sediment inflows
                if Parameter_Input_Dictionary[locs][keys]['Preferences']['File Type'] is 'xls':
                    Synthetic_Flow_Input_Data_File_Name = Parameter_Input_Dictionary[locs][keys]['Preferences']['File Path']
                    Synthetic_Flow_Input_Data_File = pd.ExcelFile(Synthetic_Flow_Input_Data_File_Name)  # Create a pandas ExcelFile object from ExcelFile class. This has performance benefits when you'll read in multiple sheets from a workbook.
                    Synthetic_Inflows_dictionary = pd.read_excel(Synthetic_Flow_Input_Data_File, sheetname=None, parse_cols=Num_Realizations-1)  # Creates a dictionary with keys = each sheet name, and content associated with each key is a dataframe created from data in each sheet. Hence, sheet names must be identical to the names of system elements in tht network connectivity matrix.
                elif Parameter_Input_Dictionary[locs][keys]['Preferences']['File Type'] is 'csv':
                    # Briefly read in each stochastic flow/sediment load file, and determine its total number of columns,
                    #  from which a random list of columns will be pulled to conduct the Monte Carlo simulation
                    col_list_dict = {}
                    for key in Parameter_Input_Dictionary[locs][keys]['Preferences']['Locations']:
                        f = Parameter_Input_Dictionary[locs][keys]['Preferences']['File Path']
                        f = open(f, 'r')
                        reader = csv.reader(f)
                        ncols = len(next(reader))  # Read first line, count columns
                        if parallelize_import is None:
                            # User not running simulation in parallel by calling start_stop parameter in pysedsim.PySedSim()
                            col_range = np.arange(ncols)
                        else:
                            # User is running stochastic simulation in parallel. Need to confine the range of columns that
                            # are pulled on this processor to its specified column range.
                            col_range = np.arange(start_stop[0], start_stop[1])

                        np.random.shuffle(col_range)  # Randomly reshuffle range
                        col_list_dict[key] = list(col_range[:Num_Realizations])
                        f.close()

                    # Read in dataframe for location. Assumes file name is same as location (which will become
                    # dictionary key)
                    keys_3 = Parameter_Input_Dictionary[locs][keys]['Preferences']['Locations'][0]
                    Synthetic_Inflows_dictionary[keys_3] = pd.read_csv(
                        Parameter_Input_Dictionary[locs][keys]['Preferences']['File Path'], nrows=len(simulation_dates),
                        usecols=col_list_dict[keys_3])

                    Synthetic_Inflows_dictionary[keys_3].columns = Parameter_Input_Dictionary[locs][keys]['Preferences']['Column Names']  # Reset column labels

                keys_3 = Parameter_Input_Dictionary[locs][keys]['Preferences']['Locations'][0]
                Synthetic_Inflows_dictionary[keys_3] = Synthetic_Inflows_dictionary[keys_3].set_index(simulation_dates)   # Loop through
                # dictionary elements (keys_3), and for each contained data frame, make the data frame have a row index equal to the dates
                #  instead of the default which is just numbers.
                Synthetic_Inflow_dataframe_name_LIST[keys_3] = Synthetic_Inflows_dictionary[keys_3].columns.values.tolist() # Also,
                # create a dictionary that stores lists of the data frames' object names
                # If flows have been specified in a form that is 365xNum_Years (leap year values were not actually generated), then account for this in the data by coping over every 2/28 value to 2/29, shifting the data each time this happens.
                if Parameter_Input_Dictionary[locs][keys]['Preferences']['Leap Year Included'] == 0:
                    simulation_dates_1 = Parameter_Input_Dictionary[locs][keys]['Preferences']['Date Range']  # Set a pandas date_range object that will serve as a dataframe index for the time series data.
                    New_Synthetic_Inflows = pd.DataFrame(index=simulation_dates_1, columns=Parameter_Input_Dictionary[locs][keys]['Preferences']['Column Names'])
                    Num_Years = simulation_dates[-1].year - simulation_dates[0].year + 1
                    start_year = simulation_dates[0].year
                    current_year = start_year
                    current_date = simulation_dates[0]  # Start date
                    num_leaps = 0
                    transfer_DF_date = current_date-timedelta(num_leaps)

                    for year in range(Num_Years):
                        if calendar.isleap(current_year) is False:
                            # Not a leap year. Slice all 365 days from inflows dictionary and store in new dataframe, then move to next year.
                            New_Synthetic_Inflows[str(current_year)] = Synthetic_Inflows_dictionary[keys_3][
                                                                       transfer_DF_date:transfer_DF_date+timedelta(364)].values
                            current_date += timedelta(365)  # proceed to first date of next year.
                        else:
                            # Leap year. Slice 31 + 28 days to cover Jan/Feb. Then copy Feb 28 value to Feb 29 value of
                            # New_Synthetic_Inflows. Then copy 29:365 from Synthetic Inflows Dictionary and and put it in 30:366 of
                            # New_Synthetic_Inflows. Then move to next year.
                            New_Synthetic_Inflows[current_date:current_date+timedelta(30+28)] = Synthetic_Inflows_dictionary[keys_3][
                                                                                                transfer_DF_date:transfer_DF_date+timedelta(30+28)].values  # Slice 31 + 28 days to cover Jan/Feb.
                            New_Synthetic_Inflows[current_date+timedelta(30+29):current_date+timedelta(30+29)] = New_Synthetic_Inflows[current_date+timedelta(30+28):current_date+timedelta(30+28)].values  # copy Feb 28 value to Feb 29 value of New_Synthetic_Inflows.
                            num_leaps += 1
                            transfer_DF_date = current_date-timedelta(num_leaps)  # track offset in New_Synthetic_Inflows compared to DF we are transferring data from.
                            New_Synthetic_Inflows[current_date+timedelta(30+30):current_date+timedelta(365)] = \
                                Synthetic_Inflows_dictionary[keys_3][transfer_DF_date+timedelta(30+30):transfer_DF_date+timedelta(
                                    365)].values  # Then copy 29:365 from Synthetic Inflows Dictionary and and put it in 30:366 of New_Synthetic_Inflows.
                            current_date += timedelta(366)  # proceed to first date of next year.
                        current_year += 1
                        transfer_DF_date = current_date-timedelta(num_leaps)  # track offset in New_Synthetic_Inflows compared to DF we are transferring data from.

                    Synthetic_Inflows_dictionary[keys_3] = New_Synthetic_Inflows
            elif keys in ['3', '5', '8', '10', '11', '12', '13', '14']:
                if Parameter_Input_Dictionary[locs][keys]['Params'][0] in ['triangular']:
                    Sampled_Parameter_Dict[locs][keys] = getattr(np.random, Parameter_Input_Dictionary[locs][keys]['Params'][0])(
                        Parameter_Input_Dictionary[locs][keys]['Params'][1], Parameter_Input_Dictionary[locs][keys]['Params'][2],
                        Parameter_Input_Dictionary[locs][keys]['Params'][3], Num_Realizations)
                elif Parameter_Input_Dictionary[locs][keys]['Params'][0] in ['uniform', 'normal', 'lognormal']:
                    Sampled_Parameter_Dict[locs][keys] = getattr(np.random, Parameter_Input_Dictionary[locs][keys]['Params'][0])(
                        Parameter_Input_Dictionary[locs][keys]['Params'][1], Parameter_Input_Dictionary[locs][keys]['Params'][2],
                        Num_Realizations)
            elif keys == '2':
                Sampled_Parameter_Dict[locs][keys] = Parameter_Input_Dictionary[locs][keys]['Params'][0]
            elif keys == '6':
                if Parameter_Input_Dictionary[locs][keys][0]['Params'][0] in ['triangular']:
                    # Creates list of numpy arrays of length Num_Realizations for the 5 parameters.
                    Sampled_Parameter_Dict[locs]['6'] = [getattr(np.random, Parameter_Input_Dictionary[locs]['6'][i]['Params'][0])(
                        Parameter_Input_Dictionary[locs]['6'][i]['Params'][1], Parameter_Input_Dictionary[locs]['6'][i]['Params'][2],
                        Parameter_Input_Dictionary[locs]['6'][i]['Params'][3], Num_Realizations) for i in range(5)]
                elif Parameter_Input_Dictionary[locs][keys][0]['Params'][0] in ['uniform', 'normal', 'lognormal']:
                    # Creates list of numpy arrays of length Num_Realizations for the 5 parameters.
                    Sampled_Parameter_Dict[locs]['6'] = [getattr(np.random, Parameter_Input_Dictionary[locs]['6'][i]['Params'][0])(
                        Parameter_Input_Dictionary[locs]['6'][i]['Params'][1], Parameter_Input_Dictionary[locs]['6'][i]['Params'][2],
                        Num_Realizations) for i in range(5)]
            elif keys == '7':
                temp_array = getattr(np.random, Parameter_Input_Dictionary[locs]['7']['Params'][0])(
                    Parameter_Input_Dictionary[locs]['7']['Params'][1], Parameter_Input_Dictionary[locs]['7']['Params'][2],
                    Num_Realizations)
                Sampled_Parameter_Dict[locs]['7'] = []  # Initialize list
                for i in range(len(temp_array)):
                    if temp_array[i] <= (1/3):
                        Sampled_Parameter_Dict[locs]['7'].append('L')
                    elif temp_array[i] <= (2/3):
                        Sampled_Parameter_Dict[locs]['7'].append('M')
                    else:
                        Sampled_Parameter_Dict[locs]['7'].append('H')
            elif keys == '14':
                if Parameter_Input_Dictionary[locs][keys][0]['Params'][0] in ['triangular']:
                    # Creates list of numpy arrays of length Num_Realizations for the 5 parameters.
                    Sampled_Parameter_Dict[locs]['6'] = [getattr(np.random, Parameter_Input_Dictionary[locs]['6'][i]['Params'][0])(
                        Parameter_Input_Dictionary[locs]['6'][i]['Params'][1], Parameter_Input_Dictionary[locs]['6'][i]['Params'][2],
                        Parameter_Input_Dictionary[locs]['6'][i]['Params'][3], Num_Realizations) for i in range(5)]
                elif Parameter_Input_Dictionary[locs][keys][0]['Params'][0] in ['uniform', 'normal', 'lognormal']:
                    # Creates list of numpy arrays of length Num_Realizations for the 5 parameters.
                    Sampled_Parameter_Dict[locs]['6'] = [getattr(np.random, Parameter_Input_Dictionary[locs]['6'][i]['Params'][0])(
                        Parameter_Input_Dictionary[locs]['6'][i]['Params'][1], Parameter_Input_Dictionary[locs]['6'][i]['Params'][2],
                        Num_Realizations) for i in range(5)]

    # Only export synthetic flows component if it exists
    if 'Synthetic_Inflow_dataframe_name_LIST' in locals() and 'Synthetic_Inflows_dictionary' in locals():
        return Sampled_Parameter_Dict, Synthetic_Inflow_dataframe_name_LIST, Synthetic_Inflows_dictionary
    else:
        # No synthetic inflows being handled; skip export
        return Sampled_Parameter_Dict

def Excel_Data_Import(element_name, Input_Data_File, Worksheet_Name, organization_type, num_data, max_distinct_data_types,
                      data_name_offset, start_date = None):
    '''
    Purpose: This module contains generalized functions that can search and import Excel workbook sheets for data
    associated with a particular name.

    :param element_name: reservoir name, dam name, channel name, etc. [string]
    :param Input_Data_File: .xlsx input file (workbook) name [string]
    :param Worksheet_Name: .xlsx worksheet within Input_Data_File workbook [string]
    :param organization_type:
        0: if time series data (column heading is element name). Only case where user needs to provide start_date.
        1: if horizontally organized data (row heading is element name)
        2: if horizontally organized stacked data
    :param num_data: The number [integer] of data points that need to be imported for the user-specified element,
    which is either a number of rows in a given column (for time series data) or a number of columns in a given row
    (for horizontally-organized data) in which data are stored for each elemeng. e.g.: Sim_Dur for a time series; 4
    for E-V-A-S; 2 for Outlet Capacity, etc.
    :param max_distinct_data_types: the number [integer] of subtypes that may exist for which there are num_data
    columns. For outlet_capacity,  this = 7, for all others, specify = None
    :param data_name_offset: number of rows between where the element name appears and where the relevant data start.
    :param start_date: date [string] in MM/DD/YYYY format, used to specify start date of time series data import.
    num_data would equal Sim_Dur (simulation duration) or Sim_Dur + 1 for time series data. If worksheet does not
    contain time series data, specify = None.
    :return:
        element_data: list or nested list of imported data
        num_outlets: if applicable, number of outlets for which data imported (integer)
    '''

    SedSim_File_Input_Sheet = Input_Data_File[Worksheet_Name]
    start_date = start_date
    if Worksheet_Name == 'Elevation Target Recurring':
        start_date = Input_Data_File['Elevation Target Recurring']['A2'].value  # Setting start date equal to value of cell A2 in "Elevation Target Recurring" worksheet ensures routine below will not search for a specific starting date, and will instead just begin at the first of 366 values.
    element_name_spacing = 1 # Spacing between element names in columns. Default = 1. e.g., Column or row names stacked on top of one another = 1.
    if data_name_offset is None:
        data_offset = 1 # This is sort of a dummy variable, used to store the value of data_name_offset specified by user as function argument. Need to be able to change the argument value in certain situations before it is used.
    else:
        data_offset = data_name_offset
    max_distinct_data_types_2 = 1  # This identicallly equal to max_distinct_data_types set by user in function argument, except it may
    # need to be changed so this new variable was established. default = 1; reset below only if max_distinct_data_types > 1
    start_offset_row = 1  # By how many rows is beginning of element name listings offset. Default value = 1
    start_offset_col = 1  # By how many rows is beginning of element name listings offset. Default value = 1

    if organization_type == 0:
        start_offset_col = 2 # By how many columns is beginning of element name listings offset.
        # Time series data
        name_row_loop = 0  # To get names, do you loop through rows? Yes = 1, No = 0
        data_row_loop = 1  # To get data, do you loop through rows? Yes = 1, No = 0
        name_col_loop = 1  # To get names, do you loop through columns? Yes = 1, No = 0
        data_col_loop = 0  # To get data, do you loop through columns? Yes = 1, No = 0
    elif organization_type == 2:
        name_row_loop = 0  # To get names, do you loop through rows? Yes = 1, No = 0
        data_row_loop = 1  # To get data, do you loop through rows? Yes = 1, No = 0
        name_col_loop = 1  # To get names, do you loop through columns? Yes = 1, No = 0
        data_col_loop = 0  # To get data, do you loop through columns? Yes = 1, No = 0
        if max_distinct_data_types is None:
            element_name_spacing = num_data # This is likely the outlet capacity sheet, where there is variable spacing between the columns. we thus cannot skip by a regular interval.
        elif max_distinct_data_types >= 1:
            max_distinct_data_types_2 = max_distinct_data_types # Value must be = 1 at a minimum
            element_name_spacing = num_data # Spacing between column names. e.g., Column names stacked on top of one another = 1.
        else:
            element_name_spacing = 1
    elif organization_type == 1:
        # Horizontal data
        name_row_loop = 1 # To get names, do you loop through rows? Yes = 1, No = 0
        data_row_loop = 0 # To get data, do you loop through rows? Yes = 1, No = 0
        name_col_loop = 0 # To get names, do you loop through columns? Yes = 1, No = 0
        data_col_loop = 1 # To get data, do you loop through columns? Yes = 1, No = 0
        start_offset_row = 2 # By how many rows is beginning of element name listings offset.
    else:
        pass

    if organization_type == 0:
        element_data = [0 for x in range(num_data)]  # Final list in which data for element will be stored.
    elif organization_type == 1:
        element_data = [None for x in range(num_data)]  # Final list in which data for element will be stored.
    elif organization_type == 2:
        if max_distinct_data_types_2 == 1:
            element_data = [[] for _ in range(num_data)] # Final list(s) in which data for element will be stored. Lists get nested if you have horizontal stacked list
        else:
            element_data = {}  # Create empty outlet dictionary = Oulet_Capacity_Dict

    exit_loop = 1
    exit_loop_2 = 1
    exit_loop_3 = 1
    counter = 0
    counter_2 = 0
    counter_4 = 0  # To keep track of how many empty columns we can loop through before we call it quits.

    while exit_loop > 0:
        if SedSim_File_Input_Sheet.cell(row = start_offset_row + counter*name_row_loop, column = start_offset_col + counter*name_col_loop*element_name_spacing).value is not None:
            counter_4 = max_distinct_data_types_2*num_data + 1  # Reset counter to check horizon of potential empty space between element names.
            if SedSim_File_Input_Sheet.cell(row = start_offset_row + counter*name_row_loop, column = start_offset_col + counter*name_col_loop*element_name_spacing).value == element_name:
                # Located user-specified element. Store data-import starting point, and load data.
                start_row = start_offset_row + counter*name_row_loop + data_offset*data_row_loop  # row starting point (where element name is located. Increment up one to get first data point.)
                start_column = start_offset_col + counter*name_col_loop*element_name_spacing + data_offset*data_col_loop  # row starting point (where element name is located. Increment up one to get first data point.)
                if organization_type == 0:
                    # If this is a time series worksheet, get correct starting and ending points based on calendar dates.
                    while exit_loop_2 > 0:
                        # Loop through time series column 1 (dates) to find point when start_date equals value in excel cell.
                        counter_2 += 1
                        if SedSim_File_Input_Sheet.cell(row = 1 + counter_2, column = 1).value != start_date:
                            pass  # Skip to next row in worksheet
                        else:
                            # Adjust starting and ending rows; columns stay the same.
                            start_row = 1 + counter_2
                            break
                else:
                    pass
                if organization_type in [0, 1]:
                    # Load data rows or columns
                    for i in range(num_data):
                        # Loop through one row or column of length num_data
                        element_data[i] = SedSim_File_Input_Sheet.cell(row = start_row + i*data_row_loop, column = start_column + i*data_col_loop).value
                elif organization_type == 2:
                    # This is a stacked horizontal type worksheet
                    if max_distinct_data_types_2 == 1:
                        for i in range(num_data):
                            # Loop through num_data columns of unknown length
                            counter_5 = 0
                            while exit_loop_3 > 0:
                                if SedSim_File_Input_Sheet.cell(row = start_row + counter_5, column = start_column + i).value is not None:
                                    element_data[i].append(SedSim_File_Input_Sheet.cell(row = start_row + counter_5, column = start_column + i).value)
                                else:
                                    break
                                counter_5 += 1
                    else:
                        # Worksheet is similar to style of Outlet Data, where there are unknown number of data types with num_data (e.g., 2) columns.
                        rowstart = start_offset_row + counter*name_row_loop
                        colstart = start_offset_col + counter*name_col_loop*element_name_spacing
                        num_outlets = int(Num_Sub_Data_Types(SedSim_File_Input_Sheet, num_data, max_distinct_data_types_2, rowstart, colstart))
                        # If reservoir was not the last for which data are stored in worksheet, loop through num_outlets for this reservoir, and store data for each outlet in an outlet dictionary.
                        for x1 in range(int(num_outlets)):
                            # Create new outlet dictionary key equal to outlet name (e.g., hydropower outlet), and store a value equal to a nested list, for outlet capacity data table to be stored in.
                            Outlet_name = SedSim_File_Input_Sheet.cell(row = rowstart + 1, column = colstart + num_data*x1).value
                            element_data[Outlet_name] = [[], []] # = Oulet_Capacity_Dict
                            # Loop through 2 columns for this outlet, and store the outlet capacity data in the nested list just created.
                            for i in range(num_data):
                                # Loop through num_data columns of unknown length
                                counter_5 = 0
                                while exit_loop_3 > 0:
                                    if SedSim_File_Input_Sheet.cell(row = start_row + counter_5, column = start_column + num_data*x1 + i).value is not None:
                                        element_data[Outlet_name][i].append(SedSim_File_Input_Sheet.cell(row = start_row + counter_5, column = start_column + num_data*x1 + i).value) # this will become Oulet_Capacity_Dict
                                    else:
                                        break
                                    counter_5 += 1
                else:
                    pass
                break # all data have now been stored, exit main while loop
            else:
                pass
        else:
            # This Excel cell contains no information (= None)
            if max_distinct_data_types_2 > 1:
                # This is likely a worksheet type similar to
                counter_4 -= 1 # decrement counter closer to zero. If it gets to zero, we have reached maximum number of loops in search of element name.
                if counter_4 == 0:
                    # print "Error 1: Element name %s missing or misspelled in workbook %s, worksheet %s. Data was not imported." % (
                    # element_name, Input_Data_File, Worksheet_Name)
                    break
                else:
                    pass
            else:
                # This is not a worksheet that has spacing between element entries, so the end of the list has been reached without locating specified element name.
                # print "Error 2: Element name %s missing or misspelled in workbook %s, worksheet %s. Data was not imported." % (
                # element_name, Input_Data_File, Worksheet_Name)
                break  # exit while loop
        counter += 1
    if 'num_outlets' in locals():
        return element_data, num_outlets
    else:
        return element_data

def Num_Sub_Data_Types(SedSim_File_Input_Sheet, num_data, max_distinct_data_types_2, rowstart, colstart):
    # Purpose: Loop to determine how many data types (e.g., outlets) each element has. This function is only called by the
    # Excel_Data_Import function (optionally).

    counter_6 = max_distinct_data_types_2*num_data + 1 # used to exit while loop that computes max # of columns in input sheet to loop over.
    counter_7 = 0
    while counter_6 > 0:
        counter_7 += 1
        if SedSim_File_Input_Sheet.cell(row = rowstart, column = colstart + counter_7).value is not None:
            num_outlets = (counter_7)/2
            break # Found another element. This is not the last element for which data are provided.
        else:
            counter_6 -=1  # decrement counter closer to 0. If it gets to 0, we have reached max # of loops in search of element name.
            if counter_6 == 0:
                # This element corresponds to the last one for which outlet data are stored. Store row number for last one.
                col_num_1 = counter_7 - (num_data*max_distinct_data_types_2 + 1)
                counter_3 = num_data
                num_outlets = 1
                exit_loop_4 = 1 # Initialize
                while exit_loop_4 > 0:
                    if SedSim_File_Input_Sheet.cell(row = rowstart + 1, column = colstart + counter_3).value is not None:
                        num_outlets += 1
                        counter_3 += num_data
                    else:
                        exit_loop_4 = 0
            else:
                pass
    return num_outlets

def Export_Simulation_Output(export_file_type, Time_Series_Output_Dictionary, state_list_excel, main_output_file_dir,
                             simulation_title, var_sub_list, rank = None, policy_name=None):

    # Purpose: To export output from simulation according to user's preferences.

    # Inputs:
    # (1) export data (Options: "Yes" or "No"). Indicates whether time series data will ultimately be exported into
    # files of specified type or not. User specifies in input file.
    # (2) export_file_type (Options: 'xls', 'xlsx', 'csv')
    # (3) Time_Series_Output_Dictionary (Created automatically from function System_Element_Creation)
    # (4) state_list_excel (Created automatically, Stores list of state var names for each object instance).
    # (9) policy_name: Name of the policy being evaluated, if pysedsim is being run in a reevaluation setting (from
    # processing_reference_set.Policy_Reevaluation()). Policy number is used as the name, and sub-folder is created.
    # Outputs:
    # None, but data are written to files and saved

    # Get operator for changing directory based on operating system.
    os_fold_op = Op_Sys_Folder_Operator()

    if rank is not None:
        cluster_sub_folder = 'cluster_output'
    else:
        cluster_sub_folder = ''

    if policy_name is not None:
        policy_folder_name = 'Policy' + '_' + str(policy_name)
    else:
        policy_folder_name = ''

    if export_file_type in ['xls', 'xlsx']:
        # This section is no longer fully functional.
        logging.info("Now beginning data export to Excel")
        # Export dataframes iteratively into excel sheets
        export_full = 0  # If = 0, then export only some objects/states dataframes; if =1 then export ALL.
        if export_full == 1:
            wb1 = {}
            writer = {}
            for keys in Time_Series_Output_Dictionary:
                # Create new workbook, name new workbook, save new workbook. Must be saved before data can be exported to workbook.
                wb1[keys] = Workbook()
                for sheetname in state_list_excel[keys]:
                    wb1[keys].create_sheet(title=sheetname)  # Load all sheet names.
                wb1[keys].save(keys + '.xlsx')  # Save file (including correctly named sheets)
                writer[keys] = pd.ExcelWriter(keys + '.xlsx', engine = 'openpyxl')
                # Export DataFrame to Excel
                for state in state_list_excel[keys]:
                    Time_Series_Output_Dictionary[keys][state].to_excel(writer[keys], state)
                    logging.info('Exporting {0} state variable time series in {1} workbook'.format(state, keys))
                writer[keys].save()
                logging.info('Saving {0} workbook'.format(keys))
        else:
            wb1 = {}
            writer = {}
            for keys in selected_objects_for_export:
                # Create new workbook, name new workbook, save new workbook. Must be saved before data can be exported to workbook.
                wb1[keys] = Workbook()
                for sheetname in selected_vars_for_export:
                    wb1[keys].create_sheet(title=sheetname)  # Load all sheet names.
                wb1[keys].save(keys + '.xlsx')  # Save file (including correctly named sheets)
                writer[keys] = pd.ExcelWriter(keys + '.xlsx', engine = 'openpyxl')
                # Export DataFrame to Excel
                for state in selected_vars_for_export:
                    Time_Series_Output_Dictionary[keys][state].to_excel(writer[keys], state)
                    logging.info('Exporting {0} state variable time series to {1} workbook'.format(state, keys))
                writer[keys].save()
                logging.info('Saving {0} workbook'.format(keys))
        logging.info("Data Export Complete")
    elif export_file_type in ['csv']:
        logging.info("Now beginning data export to .csv files")
        simulation_output_location = main_output_file_dir + os_fold_op + simulation_title + os_fold_op
        if policy_folder_name is not '':
            simulation_output_location += policy_folder_name + os_fold_op
        if cluster_sub_folder is not '':
            simulation_output_location += cluster_sub_folder + os_fold_op

        if not os.path.exists(simulation_output_location):
            os.makedirs(simulation_output_location)  # Create new folder if not existing.
        for sys_locs in Time_Series_Output_Dictionary:
            new_folder = sys_locs  # have to use two backslashes because 1 backslash is a reserved python command
            file_path = simulation_output_location + new_folder  # File path reflecting new folder
            if not os.path.exists(file_path):
                os.makedirs(file_path)  # Create new folder if it doesn't exist.
            for vars in Time_Series_Output_Dictionary[sys_locs]:
                # Loop through variables (dataframes) stored in this panel, which is associated with the key=sys_locs in the
                # Time_Series_Output_Dictionary dict. Export DF to .csv file.
                if rank is None:
                    Time_Series_Output_Dictionary[sys_locs][vars].to_csv(os.path.join(file_path, vars + '.csv'))
                else:
                    filename_processor = (vars + '_%i' + '.csv') % rank
                    Time_Series_Output_Dictionary[sys_locs][vars].to_csv(os.path.join(file_path, filename_processor))
    # Produce a README .txt file that describes each of the variables that has been exported.
    #output_variable_description(var_sub_list)
    logging.info("Finished exporting data to .csv files.")

def cluster_output_processing(n_procs, var_sub_list, simulation_title, Locations_to_Import, main_output_file_dir,
                              policies=None):
    '''

    # Purpose:This method loops through output files produced by different processors, aggregates them into a single data structure
    # (Time_Series_Output_Dictionary) and stores this single data structure as yet another .csv file according to user's preferences by
    # using the Export_Simulation_Output() function. Performance measure evaluation can then proceed as normal, wherein this csv file
    # would be imported and analysis performed.

    # Inputs:
    # (1) n_procs (Number of processors for which a batch of simulations was exported into a single csv file for each location and each
    # variable).
    # (2) export_file_type (Options: 'xsx', 'xsx', 'csv')
    # (3) state_list_excel (Created automatically, Stores list of state var names for which output files were created).

    :param policies: Optional. List of integers corresponding to row numbers of policies in the array of decision
    variable values imported from the reference set (dec_var_values, as shown below). Example: [19,3, 332]. List
    should be nested if multiple scenarios are being simulated. Example: [[4,232,1], [4, 23, 2323, 532]]

    # Outputs:
    # None, but data are written to files and saved

    # Import all relevant data for the particular scenario/system location/variable. Call Import_Simulation_Output function to loop over
    # processors.

    Args:
        n_procs:
        var_sub_list:
        simulation_title:
        Locations_to_Import:
        main_output_file_dir:
        policies:

    Returns:

    '''
    [TSID, TSID_key_list] = Import_Simulation_Output([simulation_title], Locations_to_Import, var_sub_list,
                                                     main_output_file_dir, proc_num=n_procs, policies=policies)

    # Export all aggregated data into a single file
    state_list_excel = {}  # Stores list of state var names for each object instance, truncated so name will fit into a worksheet name.
    pol_counter = 0
    for scenario in TSID_key_list:
        for loc in TSID[scenario]:
            state_list_excel[loc] = TSID[scenario][loc].keys()
            # Initialize empty list to store state
            for item in range(len(state_list_excel[loc])):
                # Only keep first 30 charac, as name will be exported to excel.
                state_list_excel[loc][item] = state_list_excel[loc][item][0:30]
        if policies is not None:
            # Feed in policy list. Then increment policy counter.
            Export_Simulation_Output('csv', TSID[scenario], state_list_excel, main_output_file_dir,
                                     simulation_title, var_sub_list, policy_name=policies[pol_counter])
            pol_counter += 1
        else:
            # User is not reevaluating policies
            Export_Simulation_Output('csv', TSID[scenario], state_list_excel, main_output_file_dir,
                                     simulation_title, var_sub_list)
    return TSID

def Import_Simulation_Output(Sims_to_Import, Locations_to_Import, var_sub_list, file_dir, proc_num = '',
                             policies=None):
    '''

    Purpose: Imports time series outputs from a simulation run(s) into a 3-dimensional pandas dataframe (DF).

    More detail: Module intended to import .csv files that have been produced as the result of a PySedSim simulation.
    Each .csv file should contain time series outputs for a single state variable (e.g., reservoir water storage) and
    system location (e.g., reservoir 1).

    The purpose of using this file may either be to (1) import the output into a dictionary of pandas structures so
    that simulation performance  measures can be evaluated and plotted, or (2) to import all the data produced by
    separate processors into a single data structure that can then be exported into a .csv file that contains aggregated
    output for each system location/variable.

    DF details (a 3D DF exists for each system location (e.g., reservoir)):
    Axis 0: State variable (e.g., Water storage, Energy production) for system location
    Axis 1: Time (e.g., dates over simulation horizon)
    Axis 2: Realization Number (e.g., stochastic simulation ensemble members)

    :param Sims_to_Import: List, containing strings of simulation scenario names (these must be directories in the
    specified output file directory that have these names). Example: ["Alternative Scenario 7A"]

    :param Locations_to_Import: Dictionary, keys store strings representing simulation element names (e.g.,
    Reservoir 1). Keys must be in the Sims_to_Import list. Example: {"Alternative Scenario 7A": ["Reservoir 1"]}

    :param var_sub_list: List, containing strings of PySedSim state variable names for which .csv output files exist
    for the scenarios in the Sims_to_Import list. Example: ['water_surface_elevation', 'capacity_active_reservoir']

    :param file_dir: String, directory in which output files to be imported are located.
    Example: r'E:\PySedSim\ModelFiles\Output_Storage'

    :param proc_num: Optional. Integer, number appended to the output .csv file representing the processor that
    produced the file (e.g., the number 3 for the file 'water_surface_elevation_3.csv')

    :param policies: Optional. List of integers corresponding to row numbers of policies in the array of decision
    variable values imported from the reference set (dec_var_values, as shown below). Example: [19,3, 332]. List
    should be nested if multiple scenarios are being simulated. Example: [[4,232,1], [4, 23, 2323, 532]]

    :return TSID: Dictionary, where keys are scenario names. Key stores sub_dictionary, where sub_dictionary keys are
    system locations storing 3D pandas DF for each system location.
    :return Num_Realizations: Dictionary, where keys are scenario names, storing number of stochastic realiztions for
    scenario.
    :return Num_Years: Dictionary, where keys are scenario names, storing number of years in a simulation realization
    for scenario
    '''

    # Import method-specific libraries
    from copy import deepcopy

    # Get operator (/ or \) for changing directory based on operating system.
    os_fold_op = Op_Sys_Folder_Operator()

    # This function reads back in previously exported simulation data so performance measure analysis can be conducted.
    if proc_num is not '':
        cluster_loop = '_0' #+ str(proc_num-1)  # Subtract 1 as first file ends with "0".
        cluster_sub_folder = 'cluster_output'
    else:
        cluster_loop = ''
        cluster_sub_folder = ''

    # Deal with situation wherein specific policies from one scenario are being reevaluated as sub-scenarios.
    master_scenario_list = []
    TSID_key_list = []
    num_scenarios = len(Sims_to_Import)
    Locations_to_Import_mod = {}
    policy_folder_name = []
    if policies is not None:
        for sim_number in range(len(Sims_to_Import)):
            if type(policies[0]) == list and len(policies) > 1 and len(policies) == num_scenarios:
                # Handle separately the case of a nested list of policy lists (for when there are multiple scenarios)
                for pol in range(len(policies[sim_number])):
                    # Store sub-scenario name and folder name
                    master_scenario_list.append(Sims_to_Import[sim_number])
                    TSID_key_list.append(Sims_to_Import[sim_number] + '-' + 'Policy ' + str(policies[sim_number][pol]))
                    policy_folder_name.append('Policy' + '_' + str(policies[sim_number][pol]))
                    Locations_to_Import_mod[Sims_to_Import[sim_number] + '-' + 'Policy ' + str(policies[sim_number][pol])] = Locations_to_Import[Sims_to_Import[sim_number]]
            else:
                for pol in range(len(policies)):
                    # Store sub-scenario name and folder name
                    master_scenario_list.append(Sims_to_Import[sim_number])
                    TSID_key_list.append(Sims_to_Import[sim_number] + '-' + 'Policy ' + str(policies[pol]))
                    policy_folder_name.append('Policy' + '_' + str(policies[pol]))
                    Locations_to_Import_mod[Sims_to_Import[sim_number] + '-' + 'Policy ' + str(policies[pol])] = \
                        Locations_to_Import[Sims_to_Import[sim_number]]
    else:
        Locations_to_Import_mod = Locations_to_Import
        for sims in Sims_to_Import:
            master_scenario_list.append(sims)
            TSID_key_list.append(sims)
            policy_folder_name.append('')

    # Initialize various data structures
    TSID = {} # Main dictionary to export
    TSID_Temp = {}  # Use to temporarily load each processor's output sheet for location/variable, if applicable
    Num_Realizations = {}  # For each scenario, stores number of realizations for that scenario
    Num_Years = {}  # For each scenario, stores number of years in each realization for that scenario
    counter = {}  # Temporary counter
    scenario_counter = 0

    # Main data import loop. Intent is to import data into Time Series Import Dictionary (TSID)
    for sims in TSID_key_list:
        counter[sims] = 0
        TSID[sims] = {}  # Sub dict for each simulation will store locations.
        TSID_Temp[sims] = {}  # Sub dict for element/variable output for a given processor in cluster.
        sim_import_loc = file_dir + os_fold_op + master_scenario_list[scenario_counter]  # This folder needs to already
        # exist.
        for sys_locs in Locations_to_Import_mod[sims]:
            TSID[sims][sys_locs] = {}  # Sub dictionary for each location will store variables.
            TSID_Temp[sims][sys_locs] = {}  # Sub dict for location will store a variable for each processor.
            if proc_num is not '':
                # Policy folders will be stored inside scenario name folder
                loc_sub_folder = os_fold_op + policy_folder_name[
                    scenario_counter] + os_fold_op + cluster_sub_folder + os_fold_op + sys_locs
            else:
                # Policy folders will be stored inside cluster_output folder.
                if policy_folder_name[scenario_counter] is not '':
                    loc_sub_folder = os_fold_op + policy_folder_name[scenario_counter] + os_fold_op + sys_locs
                else:
                    loc_sub_folder = os_fold_op + sys_locs

            # Requires that all the locs you are searching have all the variables you list above, which wont be the
            # case always (for junctions vs. reservoirs, for example).
            for vars in var_sub_list:
                file_path = sim_import_loc + loc_sub_folder  # File path reflecting new folder
                if os.path.exists(os.path.join(file_path, vars + cluster_loop + '.csv')) == True:
                    # This variable exists as a file name in the specified file path, so import it.
                    if proc_num == '':
                        # User is not importing output files produced on a cluster by various processors. Proceed
                        # linearly (there are not different files from different processors that need to be combined).
                        # Import this dataframe to a csv file.
                        TSID[sims][sys_locs][vars] = pd.read_csv(os.path.join(file_path, vars + cluster_loop + '.csv'),
                                                                 index_col=0)

                        # Force each dataframe to have datetime objects as dates rather than strings.
                        TSID[sims][sys_locs][vars].set_index(pd.to_datetime(TSID[sims][sys_locs][vars].index),
                                                             inplace=True)
                        # Determine number of realizations (ensembles). Only do this calculation once per simulation
                        # realization (on first pass through loop).
                        if counter[sims] == 0:
                            Num_Realizations[sims] = len(TSID[sims][sys_locs][vars].columns)
                            Num_Years[sims] = TSID[sims][sys_locs][vars].index[-1].year - \
                                              TSID[sims][sys_locs][vars].index[0].year + 1
                            counter[sims] += 1
                    else:
                        # User wishes to use this processor to create a dictionary for the particular
                        # location/variable of interest. This processor will therefore read in all output .csv files
                        # produced by other processors.
                        for csv_sheet in range(proc_num):
                            if os.path.exists(os.path.join(file_path, vars + '_' + str(csv_sheet) + '.csv')) == True:
                                # Import this dataframe to a csv file
                                TSID_Temp[sims][sys_locs][vars] = pd.read_csv(
                                    os.path.join(file_path, vars + '_' + str(csv_sheet) + '.csv'), index_col=0)
                                # Make each dataframe have datetime objects as dates rather than strings.
                                TSID_Temp[sims][sys_locs][vars].set_index(
                                    pd.to_datetime(TSID_Temp[sims][sys_locs][vars].index), inplace=True)
                                # Loop through locations and variables, store data from this processor in master dictionary.
                                if csv_sheet == 0:
                                    TSID[sims][sys_locs][vars] = deepcopy(TSID_Temp[sims][sys_locs][vars])
                                else:
                                    for locs in TSID[sims]:
                                        for vars in TSID[sims][locs]:
                                            # Add this new set of realizations from this DF into the main DF
                                            TSID[sims][locs][vars] = pd.concat(
                                                [TSID[sims][locs][vars], TSID_Temp[sims][locs][vars]], axis=1)
        scenario_counter += 1

    logging.info("Data Import is completed.")
    # Return is conditional. Number of realizations/years cannot be provided if the TSID only represents one of many
    # ensemble members of a stochastic simulation:
    if proc_num is not '':
        return TSID, TSID_key_list
    else:
        return TSID, Num_Realizations, Num_Years, TSID_key_list

def Total_Storage_Capacity(Sims_to_Import, Locations_to_Import, TSID):
    # This function creates a data frame for each scenario that represents reservoir total storage capacity.
    Init_total_stor = {}  # Store total storage of each simulation that's been run.
    for scenarios in Sims_to_Import:
        Init_total_stor[scenarios] = {}
        for locations in Locations_to_Import[scenarios]:
            try:
                Init_act_stor = TSID[scenarios][locations]['capacity_active_reservoir']['Realization1'].ix[0]  # Initial active storage capacity
                Init_dead_stor = TSID[scenarios][locations]['capacity_dead_reservoir']['Realization1'].ix[0] # Initial dead storage capacity
                Init_total_stor[scenarios][locations] = Init_act_stor + Init_dead_stor
                TSID[scenarios][locations]['Storage_Capacity_Total_Remaining'] = (TSID[scenarios][locations]['capacity_active_reservoir'].copy(deep=True) + TSID[scenarios][locations]['capacity_dead_reservoir'].copy(deep=True))
                TSID[scenarios][locations]['Storage_Capacity_Fract_Remaining'] = (TSID[scenarios][locations]['capacity_active_reservoir'].copy(deep=True) + TSID[scenarios][locations]['capacity_dead_reservoir'].copy(deep=True))/Init_total_stor[scenarios][locations]
                TSID[scenarios][locations]['Storage_Capacity_Loss'] = Init_total_stor[scenarios][locations] - (TSID[scenarios][locations]['capacity_active_reservoir'].copy(deep=True) + TSID[scenarios][locations]['capacity_dead_reservoir'].copy(deep=True))
                TSID[scenarios][locations]['Storage_Capacity_Loss_Percent'] = 100*(Init_total_stor[scenarios][locations] - TSID[scenarios][locations]['capacity_active_reservoir'].copy(deep=True) - TSID[scenarios][locations]['capacity_dead_reservoir'].copy(deep=True))/Init_total_stor[scenarios][locations]
            except KeyError:
                pass
    return TSID

def output_variable_description(var_sub_list):
    '''
    Purpose: to produce a text file that stores the names of variables the user has elected to produce output for,
    as well as a description of the output data being provided for each variable/.csv file.

    :param var_sub_list: A list of variable names the user desires to export during this simulation. If user did not
    specify in input, a subset of variables will automatically be included in the exported data.
    :return:
    '''

    output_variable_dict = {}  # Init dictionary

    if var_sub_list is None:
        pass  # Don't create output file storing state variable information, as no data are being exported.
    else:
        # Export file with state variable descriptions.
        for var in var_sub_list:
            output_variable_dict[var] = {}  # Init sub-dictionary
            output_variable_dict[var]['Variable Name'] = var
            output_variable_dict[var]['File Name'] = var + '.csv'

            if var == 'water_surface_elevation':
                output_variable_dict[var]['Brief Description'] = "Water surface elevation"
                output_variable_dict[var]['Units'] = 'mamsl'
                output_variable_dict[var][
                    'Extended Description'] = 'Elevation associated with the water storage in each reservoir during ' \
                                              'each time period'
            elif var == 'capacity_active_reservoir':
                output_variable_dict[var]['Brief Description'] = "Reservoir active storage capacity volume"
                output_variable_dict[var]['Units'] = 'm^3'
                output_variable_dict[var][
                    'Extended Description'] = 'Maximum capacity of a system element to store water within its ' \
                                              'active storage zone at the end of each time period. Value will not ' \
                                              'remain constant over time in a reservoir if sediment volume ' \
                                              'accumulates ' \
                                              'in the active storage zone of the reservoir.'
            elif var == 'capacity_dead_reservoir':
                output_variable_dict[var]['Brief Description'] = "Reservoir dead storage capacity volume"
                output_variable_dict[var]['Units'] = 'm^3'
                output_variable_dict[var][
                    'Extended Description'] = 'Maximum capacity of a reservoir to store water in the ' \
                                              'dead storage zone at end of each time period. Value of this variable ' \
                                              'for a particular reservoir is only different from the initial value if ' \
                                              '' \
                                              'sediment accumulates in the dead storage zone of the reservoir.'
            elif var == 'capacity_total_reservoir':
                output_variable_dict[var]['Brief Description'] = "Reservoir total storage capacity volume"
                output_variable_dict[var]['Units'] = 'm^3'
                output_variable_dict[var][
                    'Extended Description'] = 'Maximum capacity of a reservoir to store water in the value of this ' \
                                              'variable for a particular reservoir is only different from the initial ' \
                                              '' \
                                              'value if sediment accumulates in the reservoir.'
            elif var == 'Q_in':
                output_variable_dict[var]['Brief Description'] = "Daily Inflow Rate"
                output_variable_dict[var]['Units'] = 'm^3/s'
                output_variable_dict[var][
                    'Extended Description'] = 'Water discharge into a reach or reservoir during each time period.'
            elif var == 'Q_out':
                output_variable_dict[var]['Brief Description'] = 'Daily Outflow Rate'
                output_variable_dict[var]['Units'] = 'm^3/s'
                output_variable_dict[var][
                    'Extended Description'] = 'Water discharge (m^3/s) from each element (reaches and reservoirs), ' \
                                              'not including evaporation.'
            elif var == 'SS_W_in':
                output_variable_dict[var]['Brief Description'] = 'Suspended Sediment Mass Inflow'
                output_variable_dict[var]['Units'] = 'kg'
                output_variable_dict[var][
                    'Extended Description'] = 'Mass of suspended sediment that enters a system element during one ' \
                                              'time period.'
            elif var == 'SS_W_out':
                output_variable_dict[var]['Brief Description'] = 'Suspended Sediment Mass Outflow'
                output_variable_dict[var]['Units'] = 'kg'
                output_variable_dict[var]['Extended Description'] = 'Mass of suspended sediment that exits a system ' \
                                                                    'element during one time period.'
            elif var == 'BS_W':
                output_variable_dict[var]['Brief Description'] = 'Settled Sediment Mass'
                output_variable_dict[var]['Units'] = 'kg'
                output_variable_dict[var][
                    'Extended Description'] = 'Sediment mass held in bottom storage in the element ' \
                                              'of interest at the end of time period. Value can increase or ' \
                                              'decrease depending on whether scour or deposition is the dominant ' \
                                              'process.'
            elif var == 'SS_W':
                output_variable_dict[var]['Brief Description'] = 'Suspended Sediment Mass'
                output_variable_dict[var]['Units'] = 'kg'
                output_variable_dict[var][
                    'Extended Description'] = 'Mass of sediment in suspension at end of time period'
            elif var == 'TS_W':
                output_variable_dict[var]['Brief Description'] = 'Total Sediment Mass'
                output_variable_dict[var]['Units'] = 'kg'
                output_variable_dict[var][
                    'Extended Description'] = 'Sum of the bottom sediment mass and suspended sediment mass at t+1'
            elif var == 'TS_surplus_deficit':
                output_variable_dict[var][
                    'Brief Description'] = 'Total Sediment Surplus or Deficit (compared to t=0 sediment)'
                output_variable_dict[var]['Units'] = 'kg'
                output_variable_dict[var][
                    'Extended Description'] = 'Represents the total sediment that exists in a reach or reservoir ' \
                                              'at the end of each time period that is in excess (or deficit) of the ' \
                                              'amount of sediment that existed in the element at the start of ' \
                                              'simulation.'
            elif var == 'Hydropower_avg_MW':
                output_variable_dict[var]['Brief Description'] = 'Power Production'
                output_variable_dict[var]['Units'] = 'MW'
                output_variable_dict[var][
                    'Extended Description'] = 'Represents the power generated at a hydropower dam during each ' \
                                              'time period.'
            elif var == 'Hydropower_avg_MWH':
                output_variable_dict[var]['Brief Description'] = 'Energy Production'
                output_variable_dict[var]['Units'] = 'MWH'
                output_variable_dict[var][
                    'Extended Description'] = 'Represents the energy generated at a hydropower dam during each ' \
                                              'time period.'
            elif var == 'TE_avg':
                output_variable_dict[var]['Brief Description'] = 'Sediment trapping efficiency'
                output_variable_dict[var]['Units'] = 'unitless fraction'
                output_variable_dict[var][
                    'Extended Description'] = 'Represents the trap efficiency (as a fraction) for each reservoir in ' \
                                              'the system during each time period.'
            elif var == 'Residence_Time':
                output_variable_dict[var]['Brief Description'] = 'Residence time of water and sediment in a reservoir'
                output_variable_dict[var]['Units'] = 'years'
                output_variable_dict[var][
                    'Extended Description'] = 'Represents the residence time of water in each reservoir at the ' \
                                              'end of each time period.'
            elif var == 'Evap':
                output_variable_dict[var]['Brief Description'] = 'Reservoir evaporation flow rate'
                output_variable_dict[var]['Units'] = 'm^3/s'
                output_variable_dict[var][
                    'Extended Description'] = 'Represents water evaporation rate at each reservoir site during each ' \
                                              'time period.'
            elif var == 'S':
                output_variable_dict[var]['Brief Description'] = 'Water storage (volume)'
                output_variable_dict[var]['Units'] = 'm^3'
                output_variable_dict[var][
                    'Extended Description'] = 'Water volume stored in the element at the end of the time period.'
            elif var == 'Q_downstream':
                output_variable_dict[var]['Brief Description'] = 'Downstream flow rate'
                output_variable_dict[var]['Units'] = 'm^3/s'
                output_variable_dict[var][
                    'Extended Description'] = 'Water discharge released from a reservoir during each time' \
                                              'period into the element immediately downstream of the reservoir.'
            elif var == 'Q_turbines':
                output_variable_dict[var]['Brief Description'] = 'Flow rate at turbines'
                output_variable_dict[var]['Units'] = 'm^3/s'
                output_variable_dict[var][
                    'Extended Description'] = 'Water discharge released from a reservoir during each time ' \
                                              'period through the hydropower outlet(s) (turbines).'
            elif var == 'Q_spill':
                output_variable_dict[var]['Brief Description'] = 'Spilled flow rate'
                output_variable_dict[var]['Units'] = 'm^3/s'
                output_variable_dict[var][
                    'Extended Description'] = 'Water discharge rate released from a reservoir during each time ' \
                                              'period that does not generate any power.'
            elif var == 'Q_overflow':
                output_variable_dict[var]['Brief Description'] = 'Spillway flow rate'
                output_variable_dict[var]['Units'] = 'm^3/s'
                output_variable_dict[var][
                    'Extended Description'] = 'Water discharge rate released from a reservoir during each time ' \
                                              'period over the spillway outlet(s). The Q_spill variable includes this ' \
                                              '' \
                                              'flow.'
            elif var == 'Q_diversion':
                output_variable_dict[var]['Brief Description'] = 'Diversion outlet flow rate'
                output_variable_dict[var]['Units'] = 'm^3/s'
                output_variable_dict[var][
                    'Extended Description'] = 'water discharge rate released from a reservoir during each time period ' \
                                              '' \
                                              'through the diversion outlet.'
            elif var == 'Q_controlled':
                output_variable_dict[var]['Brief Description'] = 'Controlled outlet flow rate'
                output_variable_dict[var]['Units'] = 'm^3/s'
                output_variable_dict[var][
                    'Extended Description'] = 'Water discharge rate released from a reservoir during each time period ' \
                                              '' \
                                              'through the controlled outlet.'
            elif var == 'Q_low_level_outlet':
                output_variable_dict[var]['Brief Description'] = 'Low-level outlet(s) flow rate'
                output_variable_dict[var]['Units'] = 'm^3/s'
                output_variable_dict[var][
                    'Extended Description'] = 'Water discharge rate released from a reservoir during each time period ' \
                                              '' \
                                              'through the low-level outlet(s).'
            elif var == 'Q_mid_level_outlet':
                output_variable_dict[var]['Brief Description'] = 'Mid-level outlet(s) flow rate'
                output_variable_dict[var]['Units'] = 'm^3/s'
                output_variable_dict[var][
                    'Extended Description'] = 'Water discharge rate released from a reservoir during each time period ' \
                                              '' \
                                              'through the mid-level outlet(s).'
            elif var == 'Egg_Passage_Velocity_Success':
                output_variable_dict[var][
                    'Brief Description'] = 'Success/Failure to meet reservoir velocity targets for egg/larval passage'
                output_variable_dict[var]['Units'] = 'unitless binary'
                output_variable_dict[var][
                    'Extended Description'] = 'Binary evaluation (success=1/failure=0) of whether or not reservoir ' \
                                              'has ' \
                                              'met reservoir depth-averaged velocity target in time period.'
            elif var == 'Bypass_Flow_Success':
                output_variable_dict[var]['Brief Description'] = 'Bypass channel minimum flow reliability tracker'
                output_variable_dict[var]['Units'] = 'unitless binary'
                output_variable_dict[var][
                    'Extended Description'] = 'Binary evaluation (success=1/failure=0) of whether or not bypass ' \
                                              'channel' \
                                              'has met minimum flow rate target in time period. Used to compute ' \
                                              'reliability of meeting target.'
            elif var == 'bypass_flow_resilience_tracker':
                output_variable_dict[var][
                    'Brief Description'] = 'Occurrence of a failure immediately following a sucess'
                output_variable_dict[var]['Units'] = 'unitless binary'
                output_variable_dict[var][
                    'Extended Description'] = 'Binary evaluation (success=1/failure=0) of whether or not a failure to ' \
                                              '' \
                                              'meet the minimum bypass flow target immediately follows a previous ' \
                                              'failure. In this case a success indicates that a failure follows a ' \
                                              'failure.'
            elif var == 'theoretical_peaking_capacity':
                output_variable_dict[var]['Brief Description'] = 'Mid-level outlet(s) flow rate'
                output_variable_dict[var]['Units'] = 'hours'
                output_variable_dict[var][
                    'Extended Description'] = 'Number of hours/day at which power plant could operate at maximum ' \
                                              'release capacity (rated flow) while trying to empty the active storage ' \
                                              '' \
                                              'capacity. Maximum value = 24 hours.'
            elif var == 'Q_bypass':
                output_variable_dict[var]['Brief Description'] = 'Bypass channel water inflow rate'
                output_variable_dict[var]['Units'] = 'm^3/s'
                output_variable_dict[var][
                    'Extended Description'] = 'Water discharge flow rate into the bypass channel at the location of ' \
                                              'the bypass structure.'
            elif var == 'SS_W_bypass':
                output_variable_dict[var]['Brief Description'] = 'Bypass channel daily sediment inflow'
                output_variable_dict[var]['Units'] = 'kg'
                output_variable_dict[var][
                    'Extended Description'] = 'Sediment mass discharged into the bypass channel at the location of ' \
                                              'the bypass structure.'

        # Write this variable description to a text file
        with open("Time Series Output File Descriptions.txt", "w") as text_file:
            text_file.write(
                "This text file contains names and brief descriptions of time series output variables \n \n")
            for var in output_variable_dict:
                text_file.write("Variable Name: %s \n" % output_variable_dict[var]['Variable Name'])
                text_file.write("File Name: %s \n" % output_variable_dict[var]['File Name'])
                text_file.write("Brief Description: %s \n" % output_variable_dict[var]['Brief Description'])
                text_file.write("Units: %s \n" % output_variable_dict[var]['Units'])
                text_file.write("Extended Description: %s \n \n" % output_variable_dict[var]['Extended Description'])


def Import_Optimization_Preferences(simulation_title, imported_specs, main_input_files_dir):
    '''

    Purpose: To import preferences related to PySedSim reservoir operating policy optimization. Specify these
    preferences in the "Optimization" worksheet in the input data (.xlsx) file.

    Args:
        simulation_title: A list of strings of runs to conduct ("scenarios", each of which has a separate input file)
        imported_specs: A list of details regarding file locations for the scenario run(s) to be conducted,
        which represents the data contained in the top-level PySedSim input file (default file name:
        "PySedSim_Input_Specifications.csv")
        main_input_files_dir: String that contains the file path to the directory that contains the input file(s) for
        a given scenario.

    Returns:
        Borg_dict: A dictionary of Borg preferences, specified by the user in the "Optimization" worksheet of the
        pysedsim input file for the current scenario.

    '''

    Input_Data_File = Load_Input_File(simulation_title, main_input_files_dir, imported_specs)

    # Gather user-specified optimization preferences from "Optimization" worksheet if it exists.
    try:
        opt_approach = Input_Data_File['Optimization']['B2'].value
        if (opt_approach is None) or (type(opt_approach) not in [str, unicode]):
            opt_approach = 'DPS'  # Set default. Currently, only DPS is supported.

        nSeeds = Input_Data_File['Optimization']['B3'].value
        if (nSeeds is None) or (type(nSeeds) not in [int, long]):
            nSeeds = 1  # Default is 1 seeds if user did not specify anything, or didn't specify a number

        link_to_borg = Input_Data_File['Optimization']['B4'].value
        if link_to_borg in ['Yes', 'yes']:
            link_to_borg = 1
        elif link_to_borg in ['No', 'no']:
            link_to_borg = 0
        else:
            link_to_borg = 0  # Default if no user specified value
        borg_type = Input_Data_File['Optimization']['B5'].value
        if borg_type in ['Serial', 'serial']:
            borg_type = 0
        elif borg_type in ['Master-Slave', 'master-slave', 'Master Slave', 'master slave']:
            borg_type = 1
        else:
            borg_type = 0  # Default is Serial if no user specified value

        # Function evaluations
        num_func_evals = Input_Data_File['Optimization']['B6'].value
        if (num_func_evals is None) or (type(num_func_evals) not in [int, long]):
            num_func_evals = 100  # Set minimum number of function evaluations if user did not specify an integer.

        # Number of objectives
        n_objs = Input_Data_File['Optimization']['B7'].value
        if (n_objs is None) or (type(n_objs) not in [int, long]):
            n_objs = 1  # Default is a 1 objective optimization problem.
        else:
            n_objs = int(n_objs)

        n_constrs = Input_Data_File['Optimization']['B8'].value
        if (n_constrs is None) or (type(n_constrs) not in [int, long]):
            n_constrs = 0  # Default is a zero constraints in objective space
        else:
            n_constrs = int(n_constrs)

        maxtime = Input_Data_File['Optimization']['B9'].value
        if (maxtime is None) or (type(maxtime) not in [int, long]):
            maxtime = None  # Default is a zero constraints in objective space

        # Whether or not to produce reference set using Pareto.py
        ref_set = Input_Data_File['Optimization']['B10'].value
        if (ref_set is None) or (type(ref_set) not in [str, unicode]):
            ref_set = 'No'  # Default is a 1 objective optimization problem.
        else:
            if ref_set in ['yes', 'Yes', 'y', 'Y']:
                ref_set = 'Yes'
            else:
                ref_set = 'No'

        # Whether or not to produce reference set using Pareto.py, and if yes, whether or not to re-evaluate policies
        # in parallel after each optimization scenario is completed.
        re_eval = Input_Data_File['Optimization']['B11'].value
        if (re_eval is None) or (type(re_eval) not in [str, unicode]):
            re_eval = 'No'  # Default is not to automatically re-evaluate (simulate) all points in reference set.
            parallel_choice = 'No'
        else:
            if re_eval in ['yes', 'Yes', 'y', 'Y']:
                re_eval = 'Yes'
                # Whether or not to re-evaluate reference set in parallel.
                parallel_choice = Input_Data_File['Optimization']['B12'].value
                if (parallel_choice is None) or (type(parallel_choice) not in [str, unicode]):
                    parallel_choice = 'No'  # Default is not to automatically re-evaluate (simulate) all points in reference set.
                else:
                    if parallel_choice in ['yes', 'Yes', 'y', 'Y']:
                        parallel_choice = 'Yes'
                    else:
                        parallel_choice = 'No'
            else:
                re_eval = 'No'
                parallel_choice = 'No'

        # Establish user's preferences regarding plotting 2D tradeoffs in objective space, and saving plots?
        ref_set_plot = Input_Data_File['Optimization']['B13'].value
        if (ref_set_plot is None) or (type(ref_set_plot) not in [str, unicode]):
            ref_set_plot = 'No'  # Default is a 1 objective optimization problem.
        else:
            if ref_set_plot in ['yes', 'Yes', 'y', 'Y']:
                ref_set_plot = 'Yes'
            else:
                ref_set_plot = 'No'

        # Establish user's preferences regarding printing Borg runtime dynamics
        runtime_freq = Input_Data_File['Optimization']['B14'].value
        if (runtime_freq is None) or (type(runtime_freq) not in [int, long, float]):
            runtime_choice = 'No'  # Default is a 1 objective optimization problem.
        else:
            runtime_choice = 'Yes'
            runtime_freq = int(runtime_freq)
        runtime_preferences = {'runtime_choice': runtime_choice, 'runtime_freq': runtime_freq}
        # Establish preference for name of reference set file produced.
        # Currently, user cannot choose file name, but this section can be modified later, with a place added in the
        # input file for this string to be specified.

        opt_output_filename = 'pysedsim_ref_set.ref'
        obj_row = 16
        # Read in objective variable names and preferences.
        obj_var_pref = {}  # Objective variable preferences. Dictionary of variable names.
        obj_names_ordered = [0 for i in range(n_objs)]
        for i in range(n_objs):
            obj_name = Input_Data_File['Optimization'].cell(row=obj_row + i, column=1).value
            obj_names_ordered[i] = obj_name
            obj_var_pref[obj_name] = {}  # Each variable name key stores a sub dictionary of related preferences.
            obj_var_pref[obj_name]['State Variable'] = Input_Data_File['Optimization'].cell(row=obj_row + i,
                                                                                            column=2).value

            # Determine whether variable is minimization or maximization
            obj_var_pref[obj_name]['Type'] = Input_Data_File['Optimization'].cell(row=obj_row + i, column=3).value
            # Handle various ways of indicating whether variable is minimization or maximization.
            if obj_var_pref[obj_name]['Type'] not in ['Max', 'Min']:
                if obj_var_pref[obj_name]['Type'] in ['max', 'MAX', 'maximize', 'MAXIMIZE']:
                    obj_var_pref[obj_name]['Type'] = 'Max'
                elif obj_var_pref[obj_name]['Type'] in ['min', 'MIN', 'minimize', 'MINIMIZE']:
                    obj_var_pref[obj_name]['Type'] = 'Min'
                elif obj_var_pref[obj_name]['Type'] in [None]:
                    obj_var_pref[obj_name]['Type'] = 'Min'  # Assume minimization if nothing specified

            # Handle various ways of indicating what the sampling frequency in pandas should be. Options: Daily,
            # Monthly, Annual. Default: Daily.
            # For example, in a 100-year simulation, a monthly resampling would mean you draw 100*12 different
            # samples for the corresponding state variable. For each of those 1200 samples, some statistic is
            # calculated (e.g., a sum, a mean, etc.). That statistic preference is in the next section ('Resample Stat')
            obj_var_pref[obj_name]['Resample Frequency'] = Input_Data_File['Optimization'].cell(row=obj_row + i,
                                                                                                column=4).value
            if obj_var_pref[obj_name]['Resample Frequency'] not in ['A', 'M', 'D', 'O']:
                if obj_var_pref[obj_name]['Resample Frequency'] in ['Annual', 'annual', 'Annually', 'annually', 'a',
                                                                    'ann', 'Ann']:
                    obj_var_pref[obj_name]['Resample Frequency'] = 'A'
                elif obj_var_pref[obj_name]['Resample Frequency'] in ['Monthly', 'month', 'Month', 'monthly', 'm',
                                                                      'mon', 'Mon']:
                    obj_var_pref[obj_name]['Resample Frequency'] = 'M'
                elif obj_var_pref[obj_name]['Resample Frequency'] in ['Daily', 'day', 'Day', 'daily', 'd']:
                    obj_var_pref[obj_name]['Resample Frequency'] = 'D'
                elif obj_var_pref[obj_name]['Resample Frequency'] in ['Once', 'once', 'o']:
                    obj_var_pref[obj_name]['Resample Frequency'] = 'O'
                elif obj_var_pref[obj_name]['Resample Frequency'] in [None]:
                    obj_var_pref[obj_name]['Resample Frequency'] = 'D'  # Assume user is interested in mean of this
                    # variable over Monte Carlo runs.

            # Handle various ways of indicating what statistic is taken within the resampling period. For example,
            # if the resampling is Annual, and the Resample stat is Sum, then your PM is the annual sum,
            # for every year and every realization.
            # Options: Mean, Variance, Median, Max, Min, Sum. Default = Mean.
            obj_var_pref[obj_name]['Resample Stat'] = Input_Data_File['Optimization'].cell(row=obj_row + i,
                                                                                           column=5).value
            if obj_var_pref[obj_name]['Resample Stat'] not in ['mean', 'variance', 'median', 'min', 'max', 'sum']:
                if obj_var_pref[obj_name]['Resample Stat'] in ['Mean', 'MEAN', 'avg', 'AVG', 'average', 'AVERAGE']:
                    obj_var_pref[obj_name]['Resample Stat'] = 'mean'
                elif obj_var_pref[obj_name]['Resample Stat'] in ['VARIANCE', 'Variance']:
                    obj_var_pref[obj_name]['Resample Stat'] = 'Variance'
                elif obj_var_pref[obj_name]['Resample Stat'] in ['MEDIAN', 'Med']:
                    obj_var_pref[obj_name]['Resample Stat'] = 'median'
                elif obj_var_pref[obj_name]['Resample Stat'] in ['Min', 'MIN', 'minimize', 'MINIMIZE']:
                    obj_var_pref[obj_name]['Resample Stat'] = 'min'
                elif obj_var_pref[obj_name]['Resample Stat'] in ['Max', 'MAX', 'maximize', 'MAXIMIZE']:
                    obj_var_pref[obj_name]['Resample Stat'] = 'max'
                elif obj_var_pref[obj_name]['Resample Stat'] in ['SUM', 'Sum']:
                    obj_var_pref[obj_name]['Resample Stat'] = 'sum'
                elif obj_var_pref[obj_name]['Resample Stat'] in [None]:
                    obj_var_pref[obj_name]['Resample Stat'] = 'mean'  # Default

            obj_var_pref[obj_name]['Time Slice'] = Input_Data_File['Optimization'].cell(row=obj_row + i, column=6).value
            if type(obj_var_pref[obj_name]['Time Slice']) not in [str, unicode, None]:
                pass  # Slice is a number. Will determine how to use it in performance.py
            elif type(obj_var_pref[obj_name]['Time Slice']) in [str, unicode]:
                if obj_var_pref[obj_name]['Time Slice'] in ['Last', 'last']:
                    obj_var_pref[obj_name]['Time Slice'] = 1.0  # Slice at simulation duration
                else:
                    # No other strings currently supported. Set = 1 (last value).
                    obj_var_pref[obj_name]['Time Slice'] = 1.0
            else:
                obj_var_pref[obj_name]['Time Slice'] = None  # No value specified, so don't do a time slice at all.

            # Handle various ways of indicating what statistic is taken over all of the samples that are collected
            # given the specified resampling frequency and statistic. Example: 10th percentile of all monthly sums.
            # Options: Mean, Variance, Median, Max, Min, Sum, or ANY QUANTILE FRACTION [0,1]. Default = Mean.
            obj_var_pref[obj_name]['PM Distribution Stat'] = Input_Data_File['Optimization'].cell(row=obj_row + i,
                                                                                                  column=7).value
            if obj_var_pref[obj_name]['PM Distribution Stat'] not in ['mean', 'variance', 'median', 'min', 'max',
                                                                      'sum']:
                if type(obj_var_pref[obj_name]['PM Distribution Stat']) in [float, int, long]:
                    # User is specifying a quantile or percentile. If > 1 it will be assumed to be a percentage
                    # value, and will be normalized to 0-1.
                    if obj_var_pref[obj_name]['PM Distribution Stat'] <= 0 or obj_var_pref[obj_name][
                        'PM Distribution Stat'] >= 1:
                        # User has not specified a valid fraciton. Use the mean as a default.
                        obj_var_pref[obj_name]['PM Distribution Stat'] = 'mean'
                else:
                    if obj_var_pref[obj_name]['PM Distribution Stat'] in ['Mean', 'MEAN', 'avg', 'AVG', 'average',
                                                                          'AVERAGE']:
                        obj_var_pref[obj_name]['PM Distribution Stat'] = 'mean'
                    elif obj_var_pref[obj_name]['PM Distribution Stat'] in ['VARIANCE', 'Variance']:
                        obj_var_pref[obj_name]['PM Distribution Stat'] = 'Variance'
                    elif obj_var_pref[obj_name]['PM Distribution Stat'] in ['MEDIAN', 'Med']:
                        obj_var_pref[obj_name]['PM Distribution Stat'] = 'median'
                    elif obj_var_pref[obj_name]['PM Distribution Stat'] in ['Min', 'MIN', 'minimize', 'MINIMIZE']:
                        obj_var_pref[obj_name]['PM Distribution Stat'] = 'min'
                    elif obj_var_pref[obj_name]['PM Distribution Stat'] in ['Max', 'MAX', 'maximize', 'MAXIMIZE']:
                        obj_var_pref[obj_name]['PM Distribution Stat'] = 'max'
                    elif obj_var_pref[obj_name]['PM Distribution Stat'] in ['SUM', 'Sum']:
                        obj_var_pref[obj_name]['PM Distribution Stat'] = 'sum'
                    elif obj_var_pref[obj_name]['PM Distribution Stat'] in [None]:
                        obj_var_pref[obj_name]['PM Distribution Stat'] = 'mean'  # Default

            obj_var_pref[obj_name]['Epsilon'] = Input_Data_File['Optimization'].cell(row=obj_row + i, column=8).value
            if type(obj_var_pref[obj_name]['Epsilon']) not in [int, float, long]:
                obj_var_pref[obj_name]['Epsilon'] = 1  # Default epsilon value == 1

            # Read in system locations for which objective applies.
            obj_var_pref[obj_name]['Locations'] = Input_Data_File['Optimization'].cell(row=obj_row + i,
                                                                                       column=9).value.split(', ')

            obj_var_pref[obj_name]['unit_conv'] = Input_Data_File['Optimization'].cell(row=obj_row + i, column=10).value
            if type(obj_var_pref[obj_name]['unit_conv']) not in [int, float, long]:
                obj_var_pref[obj_name]['unit_conv'] = 1  # Default epsilon value == 1

            obj_var_pref[obj_name]['perc_conv'] = Input_Data_File['Optimization'].cell(row=obj_row + i, column=11).value
            if (obj_var_pref[obj_name]['perc_conv'] is None) or (
                    type(obj_var_pref[obj_name]['perc_conv']) not in [str, unicode]):
                obj_var_pref[obj_name]['perc_conv'] = 'No'  # Default is a 1 objective optimization problem.
            else:
                if obj_var_pref[obj_name]['perc_conv'] in ['yes', 'Yes', 'y', 'Y']:
                    obj_var_pref[obj_name]['perc_conv'] = 'Yes'
                else:
                    obj_var_pref[obj_name]['perc_conv'] = 'No'

            obj_var_pref[obj_name]['invert'] = Input_Data_File['Optimization'].cell(row=obj_row + i, column=12).value
            if (obj_var_pref[obj_name]['invert'] is None) or (
                    type(obj_var_pref[obj_name]['invert']) not in [str, unicode]):
                obj_var_pref[obj_name]['invert'] = 'No'  # Default is a 1 objective optimization problem.
            else:
                if obj_var_pref[obj_name]['invert'] in ['yes', 'Yes', 'y', 'Y']:
                    obj_var_pref[obj_name]['invert'] = 'Yes'
                else:
                    obj_var_pref[obj_name]['invert'] = 'No'

        # Now store object_order in the obj_var_pref dictionary
        opt_dict = {}  # Stores all data relevant to the optimization that PySedSim will need to know.
        opt_dict['Objective Names Ordered List'] = obj_names_ordered
        # Store epsilons in same order as objective names
        epsilon_list = [obj_var_pref[names]['Epsilon'] for names in opt_dict['Objective Names Ordered List']]
        opt_dict['State Variable Ordered List'] = [obj_var_pref[names]['State Variable'] for names in
                                                   opt_dict['Objective Names Ordered List']]
        opt_dict['Objective Preferences'] = obj_var_pref
        opt_dict['Num Objectives'] = n_objs
        opt_dict['Num Constraints'] = n_constrs
        opt_dict['Scenario Name'] = simulation_title
    except KeyError:
        logging.exception("Required optimization worksheet 'Optimization' does not exist in input file for "
                          "scenario {0}".format(simulation_title))

    # Read in general simulation preferences that are relevant to the optimization
    try:
        if Input_Data_File['Simulation Specifications']['B5'].value == 'Stochastic':
            Num_Realizations = Input_Data_File['Simulation Specifications']['B6'].value
        else:
            Num_Realizations = 1
    except KeyError:
        Num_Realizations = 1
    opt_dict['Num Realizations'] = Num_Realizations

    # Define Borg dictionary to return:
    Borg_dict = {
        'opt_dict': opt_dict,
        'link_to_borg': link_to_borg,
        'borg_type': borg_type,
        'num_func_evals': num_func_evals,
        'nSeeds': nSeeds,
        'n_objs': n_objs,
        'n_constrs': n_constrs,
        'maxtime': maxtime,
        'epsilon_list': epsilon_list,
        'optimization approach': opt_approach,
        'ref_set_yes_no': ref_set,
        're_eval': re_eval,
        'parallel_choice': parallel_choice,
        'opt_output_filename': opt_output_filename,
        'ref_set_plot': ref_set_plot,
        'runtime_preferences': runtime_preferences
    }

    return Borg_dict


def Op_Sys_Folder_Operator():
    '''
    Function to determine whether operating system is (1) Windows, or (2) Linux

    Returns folder operator for use in specifying directories (file locations) for reading/writing data pre- and
    post-simulation.
    '''

    if platform.system() == 'Windows':
        os_fold_op = '\\'
    elif platform.system() == 'Linux':
        os_fold_op = '/'
    else:
        os_fold_op = '/'  # Assume unix OS if it can't be identified

    return os_fold_op